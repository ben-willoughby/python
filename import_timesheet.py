# This script is one I wrote for work to import excel time entries into my outlook calendar, so I can add to the internal timesheet system easier

import pandas as pd
import win32com.client
from datetime import datetime, timedelta
from openpyxl import load_workbook
import sys

# Variables
timesheet_path = r"C:\Users\User\timesheet.xlsx"
sheet_name = "Current"
calendar_name = "Timesheet"
start_time = "09:00"
bst = True # British Summer Time
startTime_column="D"
endTime_column="E"
inCalendar_column="F"
notes_column="G"

print(calendar_name)
print(type(calendar_name))

# Connect to Outlook
outlook = win32com.client.Dispatch("Outlook.Application")
try:
    if calendar_name == "Calendar":
        calendar = outlook.GetNamespace("MAPI").GetDefaultFolder(9)
    else:
        calendar = outlook.GetNamespace("MAPI").GetDefaultFolder(9).Folders[calendar_name]
except Exception:
    print(f"Error: Calendar '{calendar_name}' not found")
    print("Check the exact name in Outlook.")
    sys.exit(1)


# read spreadsheet
df = pd.read_excel(timesheet_path)
wb = load_workbook(timesheet_path)
ws = wb[sheet_name]

# add columns if not there
if 'startTime' not in df.columns:
    df['startTime'] = ""
    ws.insert_cols(4)
if 'endTime' not in df.columns:
    df['endTime'] = ""
    ws.insert_cols(5)
if 'inCalendar' not in df.columns:
    df['inCalendar'] = ""
    ws.insert_cols(6)
if 'notes' not in df.columns:
    df['notes'] = ""
    ws.insert_cols(7)

# Format data
df['date'] = pd.to_datetime(df['date'])
df['duration'] = pd.to_numeric(df['duration'])
df['notes'] = df['notes'].fillna("").astype(str)
df['inCalendar'] = df['inCalendar'].fillna("").astype(str)
df['startTime'] = pd.to_datetime(df['startTime'], errors='coerce')
df['startTime'] = df['startTime'].dt.tz_localize(None)
df['endTime'] = pd.to_datetime(df['endTime'], errors='coerce')
df['endTime'] = df['endTime'].dt.tz_localize(None)

# create data mask for timesheet entries that are not in the calendar
not_imported_mask = df["inCalendar"].str.upper().ne("Y")

# Set start time
day_start_time = datetime.strptime(start_time, "%H:%M")
if bst:
    day_start_time = (day_start_time + timedelta(hours=1)).time()
else:
    day_start_time = day_start_time.time()

# # Add start and end times and update dataframe
for date, group in df.groupby('date'):
    first = group.index[0]

    df.loc[first, 'startTime'] = datetime.combine(
        df.loc[first, 'date'].date(),
        day_start_time
    )
    
    # add duration to starttime
    df.loc[first, 'endTime'] = df.loc[first, 'startTime'] + timedelta(minutes=int(df.loc[first, 'duration']))

    # All other rows after start
    for index in group.index[1:]:
        previous = index - 1
        df.loc[index, 'startTime'] = df.loc[previous, 'endTime']
    # add duration to start time
        df.loc[index, 'endTime'] = df.loc[index, 'startTime'] + timedelta(minutes=int(df.loc[index, 'duration']))


# Send to outlook
successful_indices = []
for index, row in df.loc[not_imported_mask].iterrows():

    appt = outlook.CreateItem(1)
    appt.Subject = row['subject']
    appt.Start = row['startTime'].tz_localize(tz='Europe/London')
    appt.End   = row['endTime'].tz_localize(tz='Europe/London')

    appt.ReminderSet = False

    appt.Move(calendar)
    successful_indices.append(index)

# Update Dataframe
importDT = datetime.now().strftime("%Y-%m-%d %H:%M")

df.loc[successful_indices, 'notes'] = f"imported via python: {importDT}"
df.loc[successful_indices, 'inCalendar'] = "Y"

# Write to excel
ws[f"{startTime_column}1"] = df['startTime'].name
ws[f"{endTime_column}1"] = df['endTime'].name
ws[f"{inCalendar_column}1"] = df['inCalendar'].name
ws[f"{notes_column}1"] = df['notes'].name

for index in successful_indices:
    excel_row = index + 2
    ws[f"{startTime_column}{excel_row}"] = df.loc[index, "startTime"]
    ws[f"{endTime_column}{excel_row}"] = df.loc[index, "endTime"]
    ws[f"{inCalendar_column}{excel_row}"] = df.loc[index, "inCalendar"]
    ws[f"{notes_column}{excel_row}"] = df.loc[index, "notes"]

wb.save(timesheet_path)